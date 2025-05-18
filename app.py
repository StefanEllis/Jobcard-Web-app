import streamlit as st
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
import io

st.set_page_config(page_title="Jobcard Auto-Filler", layout="centered")
st.title("🔧 Jobcard Auto-Filler Web App")

# Mobile-friendly style
st.markdown("""
    <style>
        input, textarea, .stTextInput, .stNumberInput, .stButton>button {
            font-size: 18px;
        }
        label {
            font-size: 16px;
        }
    </style>
""", unsafe_allow_html=True)

# Initialize session state variables
if 'continue_pressed' not in st.session_state:
    st.session_state.continue_pressed = False
if 'start_time' not in st.session_state:
    st.session_state.start_time = ""
if 'end_time' not in st.session_state:
    st.session_state.end_time = ""

use_builtin_template = st.checkbox("Use built-in jobcard template instead of uploading")

if use_builtin_template:
    uploaded_template = open("jobcard_template.xlsx", "rb")
else:
    uploaded_template = st.file_uploader("Upload your jobcard Excel template", type=["xlsx"])

email_body = st.text_area("Paste the email body here", value="", height=300, key="email_body_input")

if st.button("➡️ Continue"):
    st.session_state.continue_pressed = True

if uploaded_template and email_body and st.session_state.continue_pressed:
    def extract_fields(text):
        fields = {}
        now = datetime.now()
        fields['jobcard_number'] = 'SE2005' + now.strftime('%y%m%d%H%M')
        fields['created_date'] = now.strftime('%Y-%m-%d')

        match_wo = re.search(r"\b(WO\d{10,})\b", text)
        fields['work_order'] = match_wo.group(1) if match_wo else ""

        match_serial = re.search(r"Service Tag:\s*([A-Z0-9]{5,7})", text)
        fields['serial_number'] = match_serial.group(1) if match_serial else ""

        match_model = re.search(r"Product Model:\s*(.+)", text)
        if match_model:
            model_full = match_model.group(1).strip()
            words = model_full.split()
            if words:
                words[0] = words[0][:3].upper()
            fields['model'] = " ".join(words)
        else:
            fields['model'] = ""

        fields['csr'] = "5037"

        match_name = re.search(r"Name:\s*(.+)", text)
        fields['customer_name'] = match_name.group(1).strip() if match_name else ""

        match_contact = re.search(r"Primary Contact.*?Name:\s*(.+?)\n", text, re.DOTALL)
        fields['contact_person'] = match_contact.group(1).strip() if match_contact else ""

        match_addr = re.search(r"Customer Service Address.*?Line 1:\s*(.+?)\n.*?City:\s*(.+?)\n.*?PostalCode:\s*(\d+)", text, re.DOTALL)
        if match_addr:
            street, city, postal = match_addr.groups()
            fields['address'] = f"{street}, {city}, {postal}"
        else:
            fields['address'] = ""

        match_dell = re.search(r"\b(25\d{9})\b", text)
        fields['dell_ref'] = match_dell.group(1) if match_dell else ""

        return fields

    extracted = extract_fields(email_body)

    st.subheader("🔍 Preview & Edit Extracted Fields")
    jobcard_number = st.text_input("Jobcard Number", extracted['jobcard_number'])
    work_order = st.text_input("Work Order Number", extracted['work_order'])
    serial_number = st.text_input("Serial Number (Service Tag)", extracted['serial_number'])
    model = st.text_input("Model (First word abbreviated to 3 chars)", extracted['model'])
    csr = st.text_input("CSR Number", extracted['csr'])
    customer_name = st.text_input("Customer Name", extracted['customer_name'])
    contact_person = st.text_input("Contact Person", extracted['contact_person'])
    address = st.text_input("Customer Address", extracted['address'])
    dell_ref = st.text_input("Dell Ref (11-digit number starting with 25)", extracted['dell_ref'])
    created_date = extracted['created_date']

    kilometres = st.text_input("Kilometres Travelled")

    start_col1, start_col2 = st.columns([3, 1])
    st.session_state.start_time = start_col1.text_input("Start Time (e.g. 08:00)", st.session_state.start_time)
    if start_col2.button("-1hr"):
        st.session_state.start_time = (datetime.now() - timedelta(hours=1)).strftime("%H:%M")

    end_time_col1, end_time_col2 = st.columns([3, 1])
    st.session_state.end_time = end_time_col1.text_input("End Time (e.g. 17:00)", st.session_state.end_time)
    if end_time_col2.button("Now"):
        st.session_state.end_time = datetime.now().strftime("%H:%M")

    engineer_notes = st.text_area("Engineer Notes / Actions")

    with st.expander("➕ Add Parts (optional)"):
        num_parts = st.number_input("How many parts?", min_value=0, max_value=8, step=1, value=0, key="num_parts")
        parts = []

        for i in range(num_parts):
            st.markdown(f"**Part {i+1}**")
            part_number = st.text_input("Part Number", key=f"part_number_{i}")
            description = st.text_input("Description", key=f"part_desc_{i}")
            quantity = st.number_input("Quantity", min_value=1, step=1, value=1, key=f"part_qty_{i}")
            is_scr = st.checkbox("SCR", key=f"scr_{i}")
            parts.append((part_number, description, quantity, is_scr))

    if st.button("✅ Generate Jobcard"):
        wb = load_workbook(uploaded_template)
        ws = wb.active

        ws['D2'] = jobcard_number
        for i, char in enumerate(work_order[:12]):
            col = 7 + i
            ws.cell(row=6, column=col, value=char)

        ws['B10'] = serial_number
        ws['D10'] = model
        ws['E10'] = csr
        ws['C13'] = customer_name
        ws['C15'] = address
        ws['C17'] = contact_person
        ws['G11'] = created_date
        ws['U12'] = created_date
        ws['V12'] = st.session_state.start_time
        ws['V13'] = st.session_state.end_time
        ws['D22'] = dell_ref
        ws['P11'] = kilometres
        ws['S16'] = engineer_notes

        for i, (part_num, desc, qty, is_scr) in enumerate(parts):
            row = 22 + i
            ws[f'F{row}'] = part_num
            ws[f'L{row}'] = desc
            ws[f'V{row}'] = qty
            if is_scr:
                ws[f'T{row}'] = "(SCR)"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("🎉 Jobcard generated!")
        st.download_button(
            label="📥 Download Completed Jobcard (Excel)",
            data=output,
            file_name=f"{dell_ref}_{work_order}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
